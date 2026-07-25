"""
Importa el catálogo real de Piscilago desde data/raw/BODEGAS_Y_STOCK.xlsx
y genera data/catalog.csv (maestro de artículos, deduplicado) y
data/warehouses.csv (las 48 bodegas).

Reemplaza al antiguo scripts/generate_mock_catalog.py — ya no hay una lista
sintética que generar, este script relee la fuente real cada vez para que
el pipeline siga siendo reproducible.

Uso:
    python scripts/import_catalog.py
    python scripts/import_catalog.py --xlsx data/raw/BODEGAS_Y_STOCK.xlsx
"""
import argparse
import csv
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

STOCK_SHEETS = [
    "STOCK ALMACEN  SUMINISTROS",
    "STOCK ALMACEN AYB ",
    "STOCK RESTAURANTE FUENTES AYB",
    "STOCK RESTAURANTE FUENTES SUMIN",
    "STOCK KIOSCO TAQUILLA AYB",
    "STOCK KIOSCO PISCIGIROS AYB",
    "ZOOLOGICO",
    "ZOOLOGICO SUMINISTROS",
]
BODEGAS_SHEET = "BODEGAS DISPONIBLES"

UNIT_MAP = {
    "Unidad": "unit",
    "Kilogram": "kg",
    "Liter": "L",
    "Portion": "unit",  # sin equivalente exacto en el vocabulario UOM del sistema
}

# Orden intencional: categorías de herramientas/empaques/insumos ANTES que las
# de alimentos, porque los nombres de utensilios en español suelen incluir
# palabras de comida de forma incidental (p.ej. "MANGO" = mango de un cuchillo,
# no la fruta; "BANDEJA...HUEVOS" = una bandeja PARA huevos, no huevos;
# "BOLSA PAN..." = una bolsa PARA pan, no pan; "CUCHILLO...CORTE CARNE" = un
# cuchillo, no carne).
RULES = [
    ("Medicamentos", False, None, [
        "ACETAMINOFEN", "IBUPROFENO", "BETAMETASONA", "MG/", "MG CJ", "TAB CJ",
        "JARABE", "SHJ FCO", "AMPOLLA", "CAPSULA", "SUERO ORAL",
        "ANTIHISTAMINICO", "ANALGESICO", "DICLOFENACO", "LORATADINA",
        "LOSARTAN", "METFORMINA", "OMEPRAZOL",
    ]),
    ("Limpieza", False, None, [
        "DETERGENTE", "DESENGRASANTE", "CLOROX", "LIMPIAPISOS", "JABON",
        "DESINFECTANTE", "AMBIENTADOR", "DESMANCHADOR", "DESINCRUSTANTE",
        "BAYGON", "ANTIMICROBIANO", "ALCOHOL ANTISEPTICO",
        "ALCOHOL GLICERINADO", "GEL ANTIBACTERIAL", "CERA PISOS", "LAVALOZA",
        "LAVAPLATOS", "ESCOBA", "CEPILLO", "TRAPERO", "RECOGEDOR", "CHURRUSCO",
    ]),
    ("Desechables", False, None, [
        "BOLSA ", "SERVILLETA", "VASO DESECHABLE", "PLATO DESECHABLE",
        "PAPEL HIGIENICO", "PAPEL ABSORBENTE", "PAPEL TOALLA",
        "CUBIERTO PLASTICO", "ROLLO ", "BANDEJA EN EARTH", "BOWLS KRAFT",
        "BOWL KRAFT",
    ]),
    ("Empaques para Alimentos", False, None, [
        "CAJA PARA", "CANASTILLA", "BOLSA EMPAQUE", "BOLSA PARAFIN",
        "BOLSA PAN", "BOLSA KRAFT", "CAJA TERMOS", "CAJA MULTIFUNCIONAL",
        "BOLSA PRECORTE", "CONTENEDOR ", "COPA DESECHABLE", "COPA DE ",
    ]),
    ("Dotación y Protección Personal", False, None, [
        "BATA ", "TAPABOCAS", "GORRO DESECHABLE", "GUANTE LATEX",
        "GUANTE DE COCINA", "DELANTAL", "GUANTE DE CAUCHO",
    ]),
    ("Menaje y Cocina", False, None, [
        "CALDERO", "CAZUELA", "CACEROLA", "OLLA ", "SARTEN", "CUCHARA",
        "CUCHILLO", "TABLA PICAR", "BANDEJA", "BOWL ", "BALDE", "JARRA",
        "PLATO BLANCO", "VASO ", "CUBIERTOS", "ARAGAN", "BARRA IMANTADA",
        "BARRAIMANTADA", "ABRELATAS", "TIJERA", "ESPATULA", "PINZAS",
        "CORTADOR", "COLADOR", "ESCURRIDOR", "RALLADOR", "ESPUMADERA",
        "FRUTERA PLASTICA",
    ]),
    ("Papelería y Oficina", False, None, [
        "ARCHIVADOR", "ACETATO", "CINTA SELLAMIENTO", "MARCADOR", "LAPICERO",
        "CUADERNO", "SELLO", "ALMOHADILLA PARA SELLO", "RESMA", "GRAPA",
        "CLIPS", "FOLDER", "CARPETA", "BORRADOR PARA TABLERO", "CHINCHE",
        "ESFERO", "COSEDORA",
    ]),
    ("Señalización", False, None, [
        "BANDERA", "AVISO ", "SENAL", "VALLA", "CINTA DE PELIGRO",
        "CINTA DE SEGURIDAD",
    ]),
    ("Recreación y Zoológico", False, None, [
        "ALEVINOS", "CONCENTRADO", "PURINA", "PARA PECES", "PARA AVES",
        "COMIDA ANIMAL", "BALANCEADO", "CANOA", "FLOTADOR", "SALVAVIDAS",
        "CHALECO SALVAVIDAS", "KAYAK", "BALSA", "PISCIFLASH",
    ]),
    ("Ferretería y Mantenimiento", False, None, [
        "CANDADO", "CABO PARA", "CABUYA", "BANDA DE CAUCHO",
        "BANDAS DE CAUCHO", "ALAMBRE", "TORNILLO", "CLAVO ", "MANGUERA",
        "BOMBILLO", "EXTENSION ELECTRICA", "CINTA AISLANTE",
        "CINTA ENMASCARAR", "SILICONA", "PEGANTE", "BROCHA", "PINTURA",
        "LIJA", "DESTORNILLADOR", "MARTILLO", "LLAVE MIXTA", "TUBO PVC",
        "CODO PVC", "VALVULA", "CANDELA", "FANAL",
    ]),
    ("Almacenamiento y Organización", False, None, [
        "CAJA ORGANIZADORA", "CANASTA", "GAVETA", "ESTANTE",
        "CAJA CARTON ARCHIV", "CAJA POLICARBO", "CAJA EN CARTULINA",
        "CAJA PORTA",
    ]),
    ("Equipos y Electrónica", False, None, [
        "CALCULADORA", "LINTERNA", "PILA ", "BATERIA", "EXTENSION",
    ]),
    ("Lácteos", True, 10, [
        "LECHE", "QUESO", "YOGUR", "YOGURT", "KUMIS", "CREMA DE LECHE",
        "CUAJADA", "AREQUIPE",
    ]),
    ("Carnes", True, 3, [
        "POLLO", "CARNE MOLIDA", "CARNE", "RES ", "LOMO", "COSTILLA",
        "PECHUGA", "MUSLO", "PESCADO", "TILAPIA", "TRUCHA", "CAMARON",
        "BISTEC", "CHULETA", "CERDO", "PERNIL",
    ]),
    ("Embutidos", True, 14, [
        "SALCHICHA", "JAMON", "TOCINETA", "MORTADELA", "CHORIZO",
        "SALCHICHON", "BUTIFARRA",
    ]),
    ("Proteínas", True, 21, ["HUEVO"]),
    ("Frutas y Verduras", True, 6, [
        "TOMATE", "LECHUGA", "CEBOLLA", "PAPA ", "PAPA,", "ZANAHORIA",
        "BROCOLI", "ESPINACA", "CILANTRO", "LIMON", "NARANJA", "BANANO",
        "PINA", "MANZANA", "PERA", "UVA", "MELON", "PATILLA", "SANDIA",
        "FRESA", "MORA", "MANGO", "AGUACATE", "ACELGA", "APIO", "AHUYAMA",
        "ARANDANOS", "BERENJENA", "PEPINO", "PIMENTON", "REPOLLO", "AJO",
        "PLATANO", "YUCA", "PEREJIL", "ALBAHACA", "CALABACIN", "REMOLACHA",
        "RABANO", "ESPARRAGO", "CHAMPINON FRESCO", "CURUBA", "MARACUYA",
        "GUAYABA", "PAPAYA", "COCO", "FRIJOL VERDE", "ARVEJA VERDE FRESCA",
        "MAIZ TIERNO", "COL ",
    ]),
    ("Panadería", True, 4, [
        "PAN ", "PAN,", "CROISSANT", "AREPA", "PONQUE", "TORTA", "PASTEL",
        "BIZCOCHO",
    ]),
    ("Bebidas", False, None, [
        "AGUA ", "GASEOSA", "JUGO", "CAFE", "CHOCOLATE", "CERVEZA", "VINO",
        "LICOR", "MALTA", "REFRESCO", "ENERGIZANTE", "HIDRATANTE", "SODA",
        "AGUARDIENTE", "WHISKY", "RON ", "TE NEGRO", "TE VERDE", "LIMONADA",
        "ELECTROLIT",
    ]),
    ("Aceites", False, None, ["ACEITE", "MANTECA"]),
    ("Harinas y Cereales", False, None, [
        "HARINA", "ARROZ", "AVENA", "TRIGO", "MAIZENA", "CEREAL",
    ]),
    ("Condimentos", False, None, [
        "SAL ", "SAL,", "AZUCAR", "PANELA", "PIMIENTA", "COMINO", "VINAGRE",
        "SALSA", "SAZON", "LAUREL", "TOMILLO", "CANELA", "ACHIOTE",
        "ADEREZO", "MOSTAZA", "MAYONESA", "ESPECIA", "OREGANO",
        "CILANTRO SECO", "CONDIMENTO",
    ]),
    ("Conservas", False, None, [
        "LATA ", "ENLATAD", "ATUN", "SARDINA", "DURAZNO EN ALMIBAR",
        "COCTEL DE FRUTAS",
    ]),
    ("Pastas", False, None, ["ESPAGUETI", "MACARRON", "LASANA", "FIDEO", "PASTA "]),
    ("Abarrotes", False, None, [
        "FRIJOL", "LENTEJA", "GARBANZO", "MANI ", "GELATINA", "LEUDANTE",
        "GRANOLA", "AVENA",
    ]),
]

DEFAULT_CATEGORY = "Otros / Sin Clasificar"


def strip_diacritics(s: str) -> str:
    normalized = unicodedata.normalize("NFD", s)
    return "".join(c for c in normalized if unicodedata.category(c) != "Mn")


def classify(name: str):
    upper = strip_diacritics(name).upper()
    for cat, perish, shelf, keywords in RULES:
        for kw in keywords:
            kw_stripped = kw.rstrip()
            ends_with_space = kw != kw_stripped
            # Sin límite al final para que fragmentos de empaque médico como
            # "TAB CJ" sigan encontrando códigos fusionados como "CJX6TAB".
            pattern = r"(?<![A-Z0-9])" + re.escape(kw)
            if len(kw_stripped) <= 4 and not ends_with_space:
                # Tokens cortos (AJO, UVA...) también necesitan límite al
                # final, o coincidirían como prefijo de otra palabra más
                # larga (AJO dentro de AJONJOLI).
                pattern += r"(?![A-Z0-9])"
            if re.search(pattern, upper):
                return cat, perish, shelf
    return DEFAULT_CATEGORY, False, None


def format_code(raw) -> str | None:
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        return None
    if isinstance(raw, (int, float)):
        return str(int(round(raw)))
    raw = str(raw).strip()
    try:
        return str(int(round(float(raw))))
    except ValueError:
        return raw


def clean_name(raw) -> str:
    if raw is None:
        return ""
    return re.sub(r"\s+", " ", str(raw).strip())


def slugify(name: str) -> str:
    s = strip_diacritics(name).upper()
    s = re.sub(r"[^A-Z0-9]+", "-", s).strip("-")
    parts = [p for p in s.split("-") if p][:3]
    return "-".join(parts)


def import_warehouses(wb) -> list[dict]:
    ws = wb[BODEGAS_SHEET]
    warehouses = []
    slug_counts: dict[str, int] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Columnas: (vacío), CANTIDAD, BODEGAS
        name = clean_name(row[2] if len(row) > 2 else None)
        if not name or name == "BODEGAS":
            continue
        base = "PSL-" + slugify(name)
        slug_counts[base] = slug_counts.get(base, 0) + 1
        code = base if slug_counts[base] == 1 else f"{base}-{slug_counts[base]}"
        warehouses.append({"warehouse_code": code, "name": name})
    return warehouses


def import_catalog(wb) -> list[dict]:
    master: dict[str, dict] = {}
    no_code_counter = 0
    for sheet_name in STOCK_SHEETS:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Columnas: CANTIDAD (fila), Nr.Artículo (código), Artículo (nombre), Unidad, SD (stock — se ignora)
            if len(row) < 4:
                continue
            raw_name = row[2]
            name = clean_name(raw_name)
            if not name or name == "CANTIDAD":
                continue
            code = format_code(row[1])
            unit_raw = row[3]
            unit = UNIT_MAP.get(unit_raw, "unit")
            if not code:
                no_code_counter += 1
                code = f"SC-{no_code_counter:04d}"
            if code in master:
                continue
            category, is_perishable, shelf_days = classify(name)
            master[code] = {
                "oracle_code": code,
                "name": name,
                "unit": unit,
                "category": category,
                "is_perishable": is_perishable,
                "default_shelf_days": shelf_days,
            }
    return sorted(master.values(), key=lambda r: r["oracle_code"])


def write_csv(path: Path, rows: list[dict], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            out = dict(row)
            if "default_shelf_days" in out and out["default_shelf_days"] is None:
                out["default_shelf_days"] = ""
            writer.writerow(out)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    data_dir = Path(__file__).parent.parent / "data"
    parser.add_argument(
        "--xlsx", default=str(data_dir / "raw" / "BODEGAS_Y_STOCK.xlsx"),
        help="Ruta al archivo fuente (default: data/raw/BODEGAS_Y_STOCK.xlsx)",
    )
    args = parser.parse_args()

    wb = load_workbook(args.xlsx, data_only=True, read_only=True)

    warehouses = import_warehouses(wb)
    catalog = import_catalog(wb)

    write_csv(data_dir / "warehouses.csv", warehouses, ["warehouse_code", "name"])
    write_csv(
        data_dir / "catalog.csv",
        catalog,
        ["oracle_code", "name", "unit", "category", "is_perishable", "default_shelf_days"],
    )

    uncategorized = sum(1 for r in catalog if r["category"] == DEFAULT_CATEGORY)
    print(f"OK: {len(warehouses)} bodegas -> data/warehouses.csv")
    print(f"OK: {len(catalog)} articulos unicos -> data/catalog.csv")
    print(f"   {uncategorized} sin categoria clasificada ({DEFAULT_CATEGORY!r}) — revisar RULES si crece demasiado.")


if __name__ == "__main__":
    main()
