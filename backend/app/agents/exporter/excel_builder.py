"""Excel export con brand Colsubsidio (T-016)."""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.agents.exporter.oracle_csv import CSV_COLUMNS

HEADER_FILL = PatternFill(start_color="0067B1", end_color="0067B1", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
APPROVED_FILL = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
TITLE_FONT = Font(color="FFD000", bold=True, size=14)


def build_excel(rows: list[dict], title: str) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Conteo"

    sheet.append([title])
    sheet["A1"].font = TITLE_FONT
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(CSV_COLUMNS))

    sheet.append(CSV_COLUMNS)
    for cell in sheet[2]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for row in rows:
        sheet.append([row[col] for col in CSV_COLUMNS])
        if row.get("IS_VALIDATED") == "true":
            for cell in sheet[sheet.max_row]:
                cell.fill = APPROVED_FILL

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
